"""Excel comparison helpers."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from ..utils.filesystem import timestamped_name

HIGHLIGHT_FILL = PatternFill(
    start_color="FF0000", end_color="FF0000", fill_type="solid"
)


def highlight_differences(
    file_one: Path,
    file_two: Path,
    merge_keys: Iterable[str],
    compare_columns: Iterable[str],
    result_dir: Path,
) -> tuple[Path, Path]:
    """Highlight differences between two Excel files and write annotated copies."""

    dataframe_one = pd.read_excel(file_one)
    dataframe_two = pd.read_excel(file_two)
    merged = pd.merge(
        dataframe_one,
        dataframe_two,
        on=list(merge_keys),
        suffixes=("_file1", "_file2"),
        how="inner",
    )

    annotated_one = _annotate_workbook(file_one, merged, merge_keys, compare_columns)
    annotated_two = _annotate_workbook(file_two, merged, merge_keys, compare_columns)

    result_dir.mkdir(parents=True, exist_ok=True)

    output_one = result_dir / timestamped_name("comparison_file1", ".xlsx")
    output_two = result_dir / timestamped_name("comparison_file2", ".xlsx")

    annotated_one.save(output_one)
    annotated_two.save(output_two)

    return output_one, output_two


def _annotate_workbook(
    source_path: Path,
    merged: pd.DataFrame,
    merge_keys: Iterable[str],
    compare_columns: Iterable[str],
):
    """Apply highlighting to the workbook rows that differ."""

    workbook = load_workbook(source_path)
    worksheet = workbook.active
    key_list = list(merge_keys)
    compare_list = list(compare_columns)

    # workbook.active may be None according to some type checkers; handle that safely
    if worksheet is None:
        header_index: dict[str, int] = {}
    else:
        header_rows = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=False))
        header_row = header_rows[0] if header_rows else None
        if header_row is None:
            header_index: dict[str, int] = {}
        else:
            header_index: dict[str, int] = {
                str((cell.value) or ""): idx for idx, cell in enumerate(header_row, start=1)
            }

    # if there is no active worksheet, nothing to annotate; return early to avoid calling methods on None
    if worksheet is None:
        return workbook

    # Precompute header positions for merge keys to avoid order sensitivity
    key_positions = [header_index.get(key) for key in key_list]

    for _, row in merged.iterrows():
        key_values = tuple(row[key] for key in key_list)
        for worksheet_row in worksheet.iter_rows(min_row=2, values_only=False):
            row_key = tuple(
                worksheet_row[pos - 1].value if pos is not None and pos - 1 < len(worksheet_row) else None
                for pos in key_positions
            )
            if row_key != key_values:
                continue
            for column_name in compare_list:
                header_position = header_index.get(column_name)
                if header_position is None:
                    continue
                cell = worksheet_row[header_position - 1]
                if row.get(f"{column_name}_file1") != row.get(f"{column_name}_file2"):
                    cell.fill = HIGHLIGHT_FILL
    return workbook


def common_columns(file_one: Path, file_two: Path) -> list[str]:
    """Return the sorted common columns for the supplied Excel files."""

    dataframe_one = pd.read_excel(file_one)
    dataframe_two = pd.read_excel(file_two)
    return sorted(set(dataframe_one.columns).intersection(dataframe_two.columns))
